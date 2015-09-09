import openpyxl
import sys
import os
def main(argv,ofile):

 control= list()
 select=list()
 dos10=list()
 dos0=list()
 dos25=list()
 dos50=list()
 sum_sx=0.0
 sum_nt=0.0
 sum_ont=0.0
 for ifile in argv:
  
    wb=openpyxl.load_workbook(ifile)
    ws=wb.worksheets[0]
    for row in ws.iter_rows(row_offset=7):
       
        if row[0].internal_value!=None:
         col=(int)(row[6].internal_value)+7
         
         for c in row[col:(col+61)]:
           sum_sx+=(float)(c.internal_value)
         for c in row[col:(col+91)]:
           sum_nt+=(float)(c.internal_value)
	 for c in row[col:(col+121)]:
           sum_ont+=(float)(c.internal_value)
         if row[4].internal_value=='CON':

            control.append((row[3].internal_value,sum_sx,sum_nt,sum_ont))
         else:
 		select.append((row[3].internal_value,sum_sx,sum_nt,sum_ont))
	 if row[5].internal_value==0:
          dos0.append((row[3].internal_value,sum_sx,sum_nt,sum_ont))
	 elif row[5].internal_value==10:
            dos10.append((row[3].internal_value,sum_sx,sum_nt,sum_ont))
         elif row[5].internal_value==25:
	   dos25.append((row[3].internal_value,sum_sx,sum_nt,sum_ont))
         else:
	  dos50.append((row[3].internal_value,sum_sx,sum_nt,sum_ont))
	 sum_sx=0.0
	 sum_nt=0.0
         sum_ont=0.0
 ofile=ofile.replace('"','')
 if os.path.isfile(ofile):
    wb=openpyxl.load_workbook(ofile)
 else:
    wb=openpyxl.Workbook()
 sheet = wb.get_sheet_names()
 l=len(sheet)
 if sheet[0]== 'Activity Tracking':
  wb.remove_sheet(wb.worksheets[0])   
  ws=wb.create_sheet()
  ws.title='Activity Tracking'

 else:
                ws=wb.worksheets[0]
                ws.title='Activity Tracking'

 r=1
 ws.cell(row=r,column=1,value="Actvity information grouped by line:")
 r=r+2
 

 if control!=[]:

  ws.cell(row=r,column=1,value="Control Line:")
  r=r+2
  Title=["Animal ID","Distance travelled after 60 minutes (mm)","Distance travelled after 90 minutes (mm)","Distance travelled after 120 minutes (mm)"]
  
  for t in range(1,len(Title)+1):
     
     ws.cell(row=r,column=t,value=Title[t-1])
  r=r+1
  for c in range(1,len(control)+1):
   ws.cell(row=r,column=1,value=control[c-1][0])
   ws.cell(row=r,column=2,value=control[c-1][1])
   ws.cell(row=r,column=3,value=control[c-1][2])
   ws.cell(row=r,column=4,value=control[c-1][3])
   r=r+1
  r=r+1
 if select!=[]:
  ws.cell(row=r,column=1,value="Select Line:")
  r=r+2
  Title=["Animal ID","Distance travelled after 60 minutes (mm)","Distance travelled after 90 minutes (mm)","Distance travelled after 120 minutes (mm)"]
  
  for t in range(1,len(Title)+1):
     
     ws.cell(row=r,column=t,value=Title[t-1])
  r=r+1
  for c in range(1,len(select)+1):
   ws.cell(row=r,column=1,value=select[c-1][0])
   ws.cell(row=r,column=2,value=select[c-1][1])
   ws.cell(row=r,column=3,value=select[c-1][2])
   ws.cell(row=r,column=4,value=select[c-1][3])
   r=r+1
  r=r+1
 r=r+1
 ws.cell(row=r,column=1,value="Actvity information grouped by dose:")
 r=r+2
 
 if dos0!=[]:
    ws.cell(row=r,column=1,value="Dose 0 mg/kg")
    r=r+2
    Title=["Animal ID","Distance travelled after 60 minutes (mm)","Distance travelled after 90 minutes (mm)","Distance travelled after 120 minutes (mm)"]
  
    for t in range(1,len(Title)+1):
     
     ws.cell(row=r,column=t,value=Title[t-1])
    r=r+1
    for c in range(1,len(dos0)+1):
       ws.cell(row=r,column=1,value=dos0[c-1][0])
       ws.cell(row=r,column=2,value=dos0[c-1][1])
       ws.cell(row=r,column=3,value=dos0[c-1][2])
       ws.cell(row=r,column=4,value=dos0[c-1][3])
       r=r+1
    r=r+1
 if dos10!=[]:
    ws.cell(row=r,column=1,value="Dose 10 mg/kg")
    r=r+2
    Title=["Animal ID","Distance travelled after 60 minutes (mm)","Distance travelled after 90 minutes (mm)","Distance travelled after 120 minutes (mm)"]
   
    for t in range(1,len(Title)+1):
     
     ws.cell(row=r,column=t,value=Title[t-1])
    r=r+1
    for c in range(1,len(dos10)+1):
       ws.cell(row=r,column=1,value=dos10[c-1][0])
       ws.cell(row=r,column=2,value=dos10[c-1][1])
       ws.cell(row=r,column=3,value=dos10[c-1][2])
       ws.cell(row=r,column=4,value=dos10[c-1][3])
       r=r+1
    r=r+1
 if dos25!=[]:
    ws.cell(row=r,column=1,value="Dose 25 mg/kg")
    r=r+2
    Title=["Animal ID","Distance travelled after 60 minutes (mm)","Distance travelled after 90 minutes (mm)","Distance travelled after 120 minutes (mm)"]
   
    for t in range(1,len(Title)+1):
     
     ws.cell(row=r,column=t,value=Title[t-1])
    r=r+1
    for c in range(1,len(dos25)+1):
       ws.cell(row=r,column=1,value=dos25[c-1][0])
       ws.cell(row=r,column=2,value=dos25[c-1][1])
       ws.cell(row=r,column=3,value=dos25[c-1][2])
       ws.cell(row=r,column=4,value=dos25[c-1][3])
       r=r+1
    r=r+1
 if dos50!=[]:
    ws.cell(row=r,column=1,value="Dose 50 mg/kg")
    r=r+2
    Title=["Animal ID","Distance travelled after 60 minutes (mm)","Distance travelled after 90 minutes (mm)","Distance travelled after 120 minutes (mm)"]
    
    for t in range(1,len(Title)+1):
     
     ws.cell(row=r,column=t,value=Title[t-1])
    r=r+1
    for c in range(1,len(dos50)+1):
       ws.cell(row=r,column=1,value=dos50[c-1][0])
       ws.cell(row=r,column=2,value=dos50[c-1][1])
       ws.cell(row=r,column=3,value=dos50[c-1][2])
       ws.cell(row=r,column=4,value=dos50[c-1][3])
       r=r+1
    r=r+1

 wb.save(ofile)
if __name__=='__main__':
      ofile=raw_input('Enter output filename \t')
      ofile=ofile.strip()
      main(sys.argv[1:],ofile)
